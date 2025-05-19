import os
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
import subprocess
import sys
from datetime import datetime

# Directorio de archivos
DIRECTORIO = "C:/datosventas"
PRODUCTOS_FILE = os.path.join(DIRECTORIO, "productos.xlsx")
VENTAS_FILE = os.path.join(DIRECTORIO, "ventas.xlsx")

# Variable global para el total de la venta
total_venta = 0.0

# Crear ventana principal
root = tk.Tk()
root.title("Sistema de Ventas")

# Etiquetas
ttk.Label(root, text="Buscar producto:").grid(row=0, column=0, padx=5, pady=5)
entry_busqueda = ttk.Entry(root, width=40)
entry_busqueda.grid(row=0, column=1, padx=5, pady=5)

ttk.Label(root, text="Cantidad:").grid(row=0, column=2, padx=5, pady=5)
entry_cantidad = ttk.Entry(root, width=10)
entry_cantidad.grid(row=0, column=3, padx=5, pady=5)

# Total label
label_total = ttk.Label(root, text="Total: $0.00", font=("Arial", 12, "bold"))
label_total.grid(row=3, column=0, columnspan=5, pady=10)

# Cargar productos desde Excel
def cargar_productos():
    productos = {}
    try:
        wb = openpyxl.load_workbook(PRODUCTOS_FILE)
        hoja = wb.active
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            codigo, codigo_barra, descripcion, costo, margen, precio, stock, rubro = fila
            productos[str(codigo)] = {
                "CODIGO": str(codigo),
                "DESCRIPCION": descripcion,
                "PRECIO_DE_VENTA": float(precio),
            }
        wb.close()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar productos.\n{str(e)}")
    return productos

productos = cargar_productos()
carrito = []

# Función para volver al módulo de inicio
def volver_a_inicio(event=None):
    root.destroy()  # Cierra la ventana de ventas
    subprocess.Popen([sys.executable, "inicio.py"])  # Abre el módulo de inicio

# Actualizar total de la venta
def actualizar_total():
    global total_venta
    label_total.config(text=f"Total: ${total_venta:.2f}")

# Eliminar un producto del carrito
def eliminar_producto(event=None):
    global total_venta
    try:
        seleccion = tree.selection()
        if seleccion:
            item = tree.item(seleccion[0])
            values = item["values"]
            codigo = values[1]
            descripcion = values[2]
            cantidad = float(values[3])
            precio_unitario = float(values[4].replace("$", ""))
            total_producto = float(values[5].replace("$", ""))

            # Eliminar el producto del carrito
            for item_carrito in list(carrito):
                if (item_carrito[0] == codigo and
                    item_carrito[1] == descripcion and
                    item_carrito[2] == cantidad and
                    item_carrito[3] == precio_unitario and
                    item_carrito[4] == total_producto):
                    carrito.remove(item_carrito)
                    total_venta -= total_producto  # Restar del total_venta
                    break

            # Eliminar el producto del Treeview
            tree.delete(seleccion[0])

            # Actualizar el total
            actualizar_total()
        else:
            messagebox.showwarning("Advertencia", "Seleccione un producto para eliminar.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al eliminar el producto: {e}")

# Agregar producto al carrito
def agregar_producto(event=None):
    global producto_seleccionado, total_venta
    if producto_seleccionado:
        cantidad_str = entry_cantidad.get()
        try:
            cantidad = float(cantidad_str)
            if cantidad >= 0 and cantidad % 0.5 == 0:
                precio = producto_seleccionado["PRECIO_DE_VENTA"]
                total = precio * cantidad

                tree.insert("", tk.END, values=(
                    "",
                    producto_seleccionado["CODIGO"],
                    producto_seleccionado["DESCRIPCION"],
                    cantidad,
                    f"${precio:.2f}",
                    f"${total:.2f}"
                ))
                carrito.append((producto_seleccionado["CODIGO"], producto_seleccionado["DESCRIPCION"], cantidad, precio, total))
                # Sumar el valor del producto agregado al total de la venta
                total_venta += total
                actualizar_total()
                entry_busqueda.delete(0, tk.END)
                entry_cantidad.delete(0, tk.END)
                entry_busqueda.focus()
                # Habilitar la captura de teclas nuevamente
                root.bind("<Key>", seleccionar_producto_numero)
            else:
                messagebox.showwarning("Cantidad invalida", "Ingrese una cantidad valida (entero o fraccion de 0.5).")
        except ValueError:
            messagebox.showwarning("Cantidad invalida", "Ingrese una cantidad valida (entero o fraccion de 0.5).")
    else:
        messagebox.showwarning("Producto no seleccionado", "Seleccione primero un producto.")

# Confirmar guardado de la venta
def confirmar_guardado(metodo_pago, ventana_pago):
    global total_venta
    wb = openpyxl.load_workbook(VENTAS_FILE) if os.path.exists(VENTAS_FILE) else openpyxl.Workbook()
    hoja = wb.active

    # Obtener el último número de venta
    ultimo_numero_venta = 0
    if hoja.max_row > 1:
        for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=1, values_only=True):
            try:
                numero_venta = int(row[0])
                ultimo_numero_venta = max(ultimo_numero_venta, numero_venta)
            except (ValueError, TypeError):
                # Si no se puede convertir a entero, ignorar este valor
                pass

    if hoja.max_row == 1:
        hoja.append(["NUMERO", "CODIGO", "DESCRIPCION", "CANTIDAD", "PRECIO UNITARIO", "TOTAL", "FORMA DE PAGO", "FECHA", "HORA"])

    # Obtener la fecha y hora actual
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    hora_actual = datetime.now().strftime("%H:%M:%S")

    children = tree.get_children()
    for i, item in enumerate(carrito):
        numero_venta = ultimo_numero_venta + i + 1
        hoja.append([numero_venta] + list(item) + [metodo_pago, fecha_actual, hora_actual])
        # Actualizar el número en el Treeview solo si el índice está dentro del rango
        if i < len(children):
            tree.set(children[i], column=0, value=numero_venta)

    wb.save(VENTAS_FILE)
    wb.close()
    messagebox.showinfo("Venta guardada", "La venta ha sido guardada con éxito.")
    carrito.clear()
    tree.delete(*tree.get_children())
    total_venta = 0.0  # Restablecer el total_venta a 0.0
    actualizar_total()
    ventana_pago.destroy()

    # Establecer el foco en el campo de búsqueda para nueva venta
    entry_busqueda.focus_set()

# Guardar venta con selección de pago
def guardar_venta():
    if not carrito:
        messagebox.showwarning("Venta vacía", "No hay productos en la venta.")
        return

    # Ventana de selección de pago con botones
    ventana_pago = tk.Toplevel(root)
    ventana_pago.title("Seleccione la forma de pago")
    tk.Label(ventana_pago, text="Seleccione la forma de pago:").pack(pady=5)

    listbox_pagos = tk.Listbox(ventana_pago, height=5)
    for metodo in ["EFECTIVO", "TARJETA DE DEBITO", "TARJETA DE CREDITO", "CUENTA CORRIENTE", "TRANSFERENCIA"]:
        listbox_pagos.insert(tk.END, metodo)

    listbox_pagos.selection_set(0)  # Establecer "EFECTIVO" como opción seleccionada por defecto
    listbox_pagos.pack(pady=5)

    # Confirmar pago al seleccionar una opción
    listbox_pagos.bind("<Return>", lambda e: seleccionar_pago(listbox_pagos, ventana_pago))

    # Activar el primer elemento para navegación
    listbox_pagos.focus()

# Seleccionar forma de pago con teclado
def seleccionar_pago(listbox_pagos, ventana_pago, event=None):
    seleccion = listbox_pagos.curselection()
    if seleccion:
        index = seleccion[0]
        confirmar_guardado(listbox_pagos.get(index), ventana_pago)

# Buscar productos mientras se escribe
def buscar_productos(event):
    global producto_seleccionado, productos_mostrados
    query = entry_busqueda.get().lower()
    productos_mostrados = []
    tree_busqueda.delete(*tree_busqueda.get_children())  # Limpiar resultados anteriores
    producto_seleccionado = None

    if query:
        coincidencias = [p for p in productos.values() if query in p["DESCRIPCION"].lower()]
        for i, prod in enumerate(coincidencias):  # Mostrar todos los resultados
            tree_busqueda.insert("", tk.END, values=(i + 1, prod["DESCRIPCION"], f"${prod['PRECIO_DE_VENTA']:.2f}"))  # Agregar el precio de venta
            productos_mostrados.append(prod)
    else:
        tree_busqueda.delete(*tree_busqueda.get_children())  # Limpiar si el campo está vacío

# Seleccionar un producto de la lista (con las flechas y Enter)
def seleccionar_producto_tree(event):
    global producto_seleccionado
    try:
        seleccion = tree_busqueda.selection()
        if seleccion:
            item = tree_busqueda.item(seleccion[0])
            index = int(item["values"][0]) - 1  # Obtener el índice del producto
            producto_seleccionado = productos_mostrados[index]
            entry_busqueda.delete(0, tk.END)
            entry_busqueda.insert(0, producto_seleccionado["DESCRIPCION"])
            root.unbind("<Key>")  # Deshabilitar la captura de teclas en la ventana principal
            entry_cantidad.focus()
        else:
            messagebox.showwarning("Advertencia", "Seleccione un producto de la lista.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al seleccionar el producto: {e}")

# Seleccionar un producto de la lista (con los números)
def seleccionar_producto_numero(event):
    global producto_seleccionado
    try:
        numero_seleccionado = int(event.char)  # Obtener el número presionado
        if 1 <= numero_seleccionado <= len(productos_mostrados):
            producto_seleccionado = productos_mostrados[numero_seleccionado - 1]
            entry_busqueda.delete(0, tk.END)
            entry_busqueda.insert(0, producto_seleccionado["DESCRIPCION"])
            # Deshabilitar la captura de teclas en la ventana principal
            root.unbind("<Key>")
            entry_cantidad.focus()
            return "break"  # Consumir el evento para que no se escriba en el cuadro de búsqueda
        else:
            messagebox.showwarning("Número inválido", "Ingrese un número entre 1 y " + str(len(productos_mostrados)))
    except ValueError:
        pass  # Ignorar si no es un número

# Mover el foco al tree_busqueda al presionar Enter en el campo de búsqueda
def focus_on_tree(event):
    tree_busqueda.focus_set()
    # Seleccionar el primer elemento si existe
    items = tree_busqueda.get_children()
    if items:
        tree_busqueda.selection_set(items[0])
        tree_busqueda.focus(items[0])

# Enlazar la tecla Escape a la función volver_a_inicio
root.bind("<Escape>", volver_a_inicio)

# Botón Guardar Venta
btn_guardar = ttk.Button(root, text="Guardar Venta (F2)", command=lambda: guardar_venta())
btn_guardar.grid(row=0, column=4, padx=5, pady=5)

# Tabla de resultados de búsqueda
columns_busqueda = ("NÚMERO", "DESCRIPCIÓN", "PRECIO DE VENTA")  # Agregar la columna "PRECIO DE VENTA"
tree_busqueda = ttk.Treeview(root, columns=columns_busqueda, show="headings", height=7)  # Ajustar la altura a 7
for col in columns_busqueda:
    tree_busqueda.heading(col, text=col)
    if col == "NÚMERO":
        tree_busqueda.column(col, width=50)
    elif col == "DESCRIPCIÓN":
        tree_busqueda.column(col, width=250)
    else:
        tree_busqueda.column(col, width=100)  # Ancho para la columna "PRECIO DE VENTA"

tree_busqueda.grid(row=1, column=0, columnspan=5, padx=5, pady=5)

# Ventas table
columns = ("NÚMERO", "CÓDIGO", "DESCRIPCIÓN", "CANTIDAD", "PRECIO UNITARIO", "TOTAL")
tree = ttk.Treeview(root, columns=columns, show="headings", height=8)
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=120)

tree.grid(row=2, column=0, columnspan=5, padx=5, pady=5)

# Bind the Delete key to the eliminar_producto function
tree.bind("<Delete>", eliminar_producto)

# Bind the events and hotkeys
entry_busqueda.bind("<KeyRelease>", buscar_productos)
entry_busqueda.bind("<Return>", focus_on_tree)  # Move focus to tree_busqueda when pressing Enter
tree_busqueda.bind("<Return>", seleccionar_producto_tree)  # Select product with Enter in the tree_busqueda
root.bind("<Key>", seleccionar_producto_numero)  # Capture the keys pressed in the main window (for numbers)
entry_cantidad.bind("<Return>", agregar_producto)
root.bind("<F2>", lambda e: guardar_venta())  # Bind the F2 key to the guardar_venta function

# Set focus on the search field at startup
entry_busqueda.focus_set()

actualizar_total()

root.mainloop()
